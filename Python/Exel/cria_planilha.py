
# No VScode vale a pena usar a extensão Excel View

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# Este comando usa uma planilha já ativa:
# worksheet: Worksheet = workbook.active

name_planilha = 'Personagens'
# Cria e coloca como primeira na ordem seleção
workbook.create_sheet(name_planilha, 0)
# Seleciona a planilha
worksheet: Worksheet = workbook[name_planilha]

# Escolhe linha e coluna para cada name.
worksheet.cell(1, 1, 'Name')
worksheet.cell(1, 2, 'Age')
worksheet.cell(1, 3, 'Class')


players = [
    ['TioZiio', 23, 'God'],
    ['Arthur Morgan', 16, 'Diplomata'],
    ['Percival', 16, 'Guerreiro'],
]

# Método diferente, mais fácil
for n in players:
    worksheet.append(n)

# Método com Matriz
# for i, n in enumerate(players, start=2):
#     for j, k in enumerate(n, start=1):
#         worksheet.cell(i, j, k)

workbook.save(WORKBOOK_PATH)
