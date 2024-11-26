
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook: Workbook = load_workbook(WORKBOOK_PATH)
worksheet: Worksheet = workbook['Personagens']

n : tuple[Cell]
for n in worksheet.iter_rows(min_row=2):
    # Tupla de células : <Cell 'Personagens'.A2>
    for cell in n:
        print(cell.value, end='\t')

        if cell.value == 'Percival':
            worksheet.cell(cell.row, 1, 'Perci')
    print()

# Selecionando valor específico
print(worksheet['A3'].value)
# Alterando valor específico
worksheet['A3'].value = 'Arthur'

workbook.save(WORKBOOK_PATH)
